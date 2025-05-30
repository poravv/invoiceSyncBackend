import os
import logging
import pandas as pd
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.models.models import InvoiceData
from app.config.settings import settings

logger = logging.getLogger(__name__)

class ExcelExporter:
    def __init__(self, output_path: str = None):
        """
        Inicializa el exportador a Excel.
        
        Args:
            output_path: Ruta del archivo Excel de salida. Si no se proporciona,
                       se utiliza el valor de configuración.
        """
        self.output_path = output_path or settings.EXCEL_OUTPUT_PATH
        
        # Crear directorio si no existe
        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
    
    def export_invoices(self, invoices: List[InvoiceData]) -> str:
        """
        Exporta facturas a un archivo Excel con todos los campos necesarios.
        
        Args:
            invoices: Lista de objetos InvoiceData para exportar.
            
        Returns:
            str: Ruta del archivo Excel generado.
        """
        if not invoices:
            logger.warning("No hay facturas para exportar")
            return ""
        
        try:
            # Convertir a lista de diccionarios para pandas
            data = []
            productos_data = []  # Para la hoja de productos
            
            for i, invoice in enumerate(invoices):
                # Convertir fecha a str para mejor visualización
                fecha_str = invoice.fecha.strftime("%d/%m/%Y") if invoice.fecha else ""
                procesado_str = invoice.procesado_en.strftime("%d/%m/%Y %H:%M:%S") if invoice.procesado_en else ""
                
                # Extraer información de empresa, si está disponible
                empresa_nombre = ""
                empresa_direccion = ""
                empresa_telefono = ""
                
                if invoice.empresa:
                    if isinstance(invoice.empresa, dict):
                        empresa_nombre = invoice.empresa.get('nombre', '')
                        empresa_direccion = invoice.empresa.get('direccion', '')
                        empresa_telefono = invoice.empresa.get('telefono', '')
                    else:
                        empresa_nombre = invoice.empresa.nombre if hasattr(invoice.empresa, 'nombre') else ""
                        empresa_direccion = invoice.empresa.direccion if hasattr(invoice.empresa, 'direccion') else ""
                        empresa_telefono = invoice.empresa.telefono if hasattr(invoice.empresa, 'telefono') else ""
                
                # Extraer información del timbrado
                timbrado_inicio = ""
                timbrado_fin = ""
                
                if invoice.timbrado_data:
                    if isinstance(invoice.timbrado_data, dict):
                        timbrado_inicio = invoice.timbrado_data.get('fecha_inicio_vigencia', '')
                        timbrado_fin = invoice.timbrado_data.get('valido_hasta', '')
                    else:
                        timbrado_inicio = invoice.timbrado_data.fecha_inicio_vigencia if hasattr(invoice.timbrado_data, 'fecha_inicio_vigencia') else ""
                        timbrado_fin = invoice.timbrado_data.valido_hasta if hasattr(invoice.timbrado_data, 'valido_hasta') else ""
                
                # Información de totales
                total_iva = 0
                subtotal = 0
                
                if invoice.totales:
                    if isinstance(invoice.totales, dict):
                        total_iva = float(invoice.totales.get('total_iva', 0))
                        subtotal = float(invoice.totales.get('subtotal', 0))
                    else:
                        total_iva = float(invoice.totales.total_iva) if hasattr(invoice.totales, 'total_iva') else 0
                        subtotal = float(invoice.totales.subtotal) if hasattr(invoice.totales, 'subtotal') else 0
                
                # Datos principales para la hoja principal
                data.append({
                    "Fecha": fecha_str,
                    "RUC Emisor": invoice.ruc_emisor or "",
                    "Nombre Emisor": invoice.nombre_emisor or empresa_nombre,
                    "Dirección Emisor": empresa_direccion,
                    "Teléfono Emisor": empresa_telefono,
                    "Nro. Factura": invoice.numero_factura or "",
                    "Condición Venta": invoice.condicion_venta or "",
                    "Moneda": invoice.moneda or "PYG",
                    "Monto Total": float(invoice.monto_total) if invoice.monto_total else 0.0,
                    "Subtotal": subtotal,
                    "IVA": float(invoice.iva) if invoice.iva else total_iva,
                    "Subtotal Exentas": float(invoice.subtotal_exentas) if invoice.subtotal_exentas else 0.0,
                    "Subtotal 5%": float(invoice.subtotal_5) if invoice.subtotal_5 else 0.0,
                    "Subtotal 10%": float(invoice.subtotal_10) if invoice.subtotal_10 else 0.0,
                    "RUC Cliente": invoice.ruc_cliente or "",
                    "Nombre Cliente": invoice.nombre_cliente or "",
                    "Email Cliente": invoice.email_cliente or "",
                    "Timbrado": invoice.timbrado or "",
                    "Timbrado Inicio": timbrado_inicio,
                    "Timbrado Fin": timbrado_fin,
                    "CDC": invoice.cdc or "",
                    "Actividad Económica": invoice.actividad_economica or "",
                    "Productos": len(invoice.productos) if invoice.productos else 0,
                    "PDF": invoice.pdf_path or "",
                    "Origen (correo)": invoice.email_origen or "",
                    "Procesado en": procesado_str
                })
                
                # Añadir datos de productos
                if invoice.productos:
                    for producto in invoice.productos:
                        # Manejar productos tanto como objetos ProductoFactura como diccionarios
                        if isinstance(producto, dict):
                            articulo = producto.get('articulo', '')
                            cantidad = float(producto.get('cantidad', 0))
                            precio_unitario = float(producto.get('precio_unitario', 0))
                            total = float(producto.get('total', 0))
                        else:
                            articulo = producto.articulo if hasattr(producto, 'articulo') else ""
                            cantidad = float(producto.cantidad) if hasattr(producto, 'cantidad') else 0
                            precio_unitario = float(producto.precio_unitario) if hasattr(producto, 'precio_unitario') else 0
                            total = float(producto.total) if hasattr(producto, 'total') else 0
                        
                        productos_data.append({
                            "Factura": invoice.numero_factura or "",
                            "RUC Emisor": invoice.ruc_emisor or "",
                            "Fecha": fecha_str,
                            "Artículo": articulo,
                            "Cantidad": cantidad,
                            "Precio Unitario": precio_unitario,
                            "Total": total
                        })
            
            # Columnas numéricas para formateo especial
            numeric_cols = ["Monto Total", "Subtotal", "IVA", "Subtotal Exentas", "Subtotal 5%", "Subtotal 10%", "Productos"]
            productos_numeric_cols = ["Cantidad", "Precio Unitario", "Total"]
            
            # Crear o cargar el archivo Excel existente
            if os.path.exists(self.output_path):
                try:
                    existing_df = pd.read_excel(self.output_path, sheet_name="Facturas")
                    new_df = pd.DataFrame(data)
                    
                    # Convertir columnas numéricas a float para evitar problemas
                    for col in numeric_cols:
                        if col in existing_df.columns and col in new_df.columns:
                            existing_df[col] = existing_df[col].astype(float)
                            new_df[col] = new_df[col].astype(float)
                    
                    # Concatenar y eliminar duplicados
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                    combined_df.drop_duplicates(
                        subset=["RUC Emisor", "Nro. Factura", "Monto Total", "CDC"],
                        keep="last",
                        inplace=True
                    )
                    
                    # Cargar datos de productos existentes
                    try:
                        existing_productos_df = pd.read_excel(self.output_path, sheet_name="Productos")
                        productos_df = pd.DataFrame(productos_data)
                        
                        # Convertir columnas numéricas de productos a float
                        for col in productos_numeric_cols:
                            if col in existing_productos_df.columns and col in productos_df.columns:
                                existing_productos_df[col] = existing_productos_df[col].astype(float)
                                productos_df[col] = productos_df[col].astype(float)
                        
                        # Combinar productos
                        combined_productos_df = pd.concat([existing_productos_df, productos_df], ignore_index=True)
                        combined_productos_df.drop_duplicates(
                            subset=["Factura", "RUC Emisor", "Artículo"],
                            keep="last",
                            inplace=True
                        )
                    except Exception as e:
                        logger.warning(f"No se encontró hoja de productos existente: {str(e)}")
                        combined_productos_df = pd.DataFrame(productos_data)
                    
                    # Guardar el DataFrame combinado en múltiples hojas
                    with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                        combined_df.to_excel(writer, sheet_name="Facturas", index=False)
                        combined_productos_df.to_excel(writer, sheet_name="Productos", index=False)
                    
                except Exception as e:
                    logger.error(f"Error al cargar archivo Excel existente: {str(e)}")
                    # Si hay error al cargar, creamos uno nuevo
                    with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                        pd.DataFrame(data).to_excel(writer, sheet_name="Facturas", index=False)
                        pd.DataFrame(productos_data).to_excel(writer, sheet_name="Productos", index=False)
            else:
                # Crear nuevo archivo con múltiples hojas
                with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                    pd.DataFrame(data).to_excel(writer, sheet_name="Facturas", index=False)
                    pd.DataFrame(productos_data).to_excel(writer, sheet_name="Productos", index=False)
            
            # Aplicar formato
            self._apply_excel_formatting(numeric_cols, productos_numeric_cols)
            
            logger.info(f"Archivo Excel generado: {self.output_path} con {len(data)} facturas")
            return self.output_path
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {str(e)}", exc_info=True)
            return ""
    
    def _apply_excel_formatting(self, numeric_cols: List[str], productos_numeric_cols: List[str]):
        """
        Aplica formato al archivo Excel para mejorar su visualización.
        
        Args:
            numeric_cols: Lista de columnas numéricas para formateo especial en la hoja principal.
            productos_numeric_cols: Lista de columnas numéricas para formateo especial en la hoja de productos.
        """
        try:
            wb = openpyxl.load_workbook(self.output_path)
            
            # Definir estilos
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Formatear cada hoja
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Determinar qué columnas numéricas usar según la hoja
                sheet_numeric_cols = numeric_cols if sheet_name == "Facturas" else productos_numeric_cols
                
                # Aplicar formato a encabezados
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Aplicar formato a celdas
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")
                        
                        # Formato para columnas numéricas
                        header_value = ws.cell(1, cell.column).value
                        if header_value in sheet_numeric_cols:
                            cell.number_format = "#,##0.00"
                
                # Auto-ajustar anchos de columna
                for column in ws.columns:
                    max_length = max(
                        len(str(cell.value)) for cell in column
                    ) if any(cell.value for cell in column) else 0
                    
                    if max_length > 0:
                        column_letter = column[0].column_letter
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                # Congelar paneles (fila de encabezados)
                ws.freeze_panes = "A2"
            
            # Guardar cambios
            wb.save(self.output_path)
            logger.info("Formato aplicado al archivo Excel")
            
        except Exception as e:
            logger.error(f"Error al aplicar formato Excel: {str(e)}", exc_info=True)
    
    def append_invoices(self, invoices: List[InvoiceData]) -> bool:
        """
        Añade facturas al archivo Excel existente.
        
        Args:
            invoices: Lista de objetos InvoiceData para añadir.
            
        Returns:
            bool: True si se añadieron correctamente, False en caso contrario.
        """
        return bool(self.export_invoices(invoices))